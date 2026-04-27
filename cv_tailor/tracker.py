"""Job application tracker — appends rows to job_applications.xlsx."""
import csv
import os
import threading
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TRACKER_FILENAME = "job_applications.xlsx"

HEADERS = [
    "#", "Date Applied", "Company", "Role", "Fit", "Fit Score",
    "Fit Summary", "Hard Gap", "CV Filename", "Status",
    "Interview Date", "Notes", "Follow-up Due",
]

COL_WIDTHS = {
    "A": 5, "B": 14, "C": 22, "D": 30, "E": 10, "F": 10, "G": 35,
    "H": 30, "I": 40, "J": 18, "K": 14, "L": 30, "M": 14,
}

STATUS_OPTIONS = [
    "Applied", "Viewed", "Phone Screen", "Interview Scheduled",
    "Interviewed", "Technical Test", "Offer", "Rejected",
    "Ghosted", "Withdrawn",
]

ACTIVE_EXCLUDED = {"Rejected", "Ghosted", "Withdrawn"}

HEADER_FILL = PatternFill("solid", fgColor="0d1117")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
ROW_FONT = Font(name="Calibri", size=10)
ALT_FILL = PatternFill("solid", fgColor="f6f8fa")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

FIT_STYLES = {
    "Green": (PatternFill("solid", fgColor="d4edda"),
              Font(name="Calibri", size=10, bold=True, color="2ea44f")),
    "Yellow": (PatternFill("solid", fgColor="fff3cd"),
               Font(name="Calibri", size=10, bold=True, color="d29922")),
    "Red": (PatternFill("solid", fgColor="ffdce0"),
            Font(name="Calibri", size=10, bold=True, color="da3633")),
}

tracker_lock = threading.Lock()


def get_tracker_path(output_folder):
    """Return absolute path to the tracker xlsx in the given output folder."""
    return os.path.join(output_folder, TRACKER_FILENAME)


def _format_date(dt=None):
    """Return today's date as DD MMM YYYY."""
    return (dt or datetime.now()).strftime("%d %b %Y")


def _normalize_fit(fit_value):
    """Map a fit indicator (str/icon) to canonical Green/Yellow/Red."""
    if not fit_value:
        return ""
    s = str(fit_value).lower()
    if "green" in s or "strong" in s:
        return "Green"
    if "yellow" in s or "partial" in s:
        return "Yellow"
    if "red" in s or "poor" in s:
        return "Red"
    return ""


def _create_workbook(path):
    """Create a fresh tracker workbook with header row and formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 30
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")

    ws.freeze_panes = "A2"
    _attach_status_validation(ws)
    wb.save(path)
    return wb


def _attach_status_validation(ws):
    """Attach the Status dropdown to column J for many rows."""
    formula = '"' + ",".join(STATUS_OPTIONS) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add("J2:J1048576")
    ws.add_data_validation(dv)


def _load_or_create(path):
    """Load existing tracker workbook or create a new one."""
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception:
            pass
    return _create_workbook(path)


def _apply_row_styling(ws, row_idx, fit_value):
    """Apply alternating fill, font, height, and Fit cell styling."""
    fill = ALT_FILL if (row_idx % 2 == 0) else WHITE_FILL
    ws.row_dimensions[row_idx].height = 18
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = ROW_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=False)

    canonical = _normalize_fit(fit_value)
    if canonical in FIT_STYLES:
        fit_fill, fit_font = FIT_STYLES[canonical]
        fit_cell = ws.cell(row=row_idx, column=5)
        fit_cell.value = canonical
        fit_cell.fill = fit_fill
        fit_cell.font = fit_font
        fit_cell.alignment = Alignment(vertical="center", horizontal="center")


def find_existing_application(output_folder, company, role):
    """Return (row_idx, date_str) for first existing match, or None."""
    path = get_tracker_path(output_folder)
    if not os.path.exists(path):
        return None
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            r_company = (row[2] or "").strip() if row[2] else ""
            r_role = (row[3] or "").strip() if row[3] else ""
            if (r_company.lower() == (company or "").strip().lower()
                    and r_role.lower() == (role or "").strip().lower()):
                return (r_company, r_role, row[1] or "")
    except Exception:
        return None
    return None


def write_tracker_row(output_folder, row_data):
    """Append a single row to the job application tracker.

    row_data keys: company, role, fit, fit_score, fit_summary, hard_gap,
    cv_filename. Date is generated here.
    """
    os.makedirs(output_folder, exist_ok=True)
    path = get_tracker_path(output_folder)
    with tracker_lock:
        wb = _load_or_create(path)
        ws = wb.active
        next_num = ws.max_row  # header is row 1, so existing data rows = max_row-1; new # = max_row
        new_row_idx = ws.max_row + 1
        fit_canonical = _normalize_fit(row_data.get("fit", ""))

        values = [
            next_num,
            _format_date(),
            row_data.get("company", ""),
            row_data.get("role", ""),
            fit_canonical,
            row_data.get("fit_score", ""),
            row_data.get("fit_summary", ""),
            row_data.get("hard_gap", ""),
            row_data.get("cv_filename", ""),
            "", "", "", "",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=new_row_idx, column=col_idx, value=val)

        _apply_row_styling(ws, new_row_idx, fit_canonical)
        wb.save(path)


def read_all_rows(output_folder):
    """Return list of dicts (one per data row) from the tracker."""
    path = get_tracker_path(output_folder)
    rows = []
    if not os.path.exists(path):
        return rows
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            padded = list(row) + [""] * (len(HEADERS) - len(row))
            rec = {HEADERS[i]: ("" if padded[i] is None else padded[i])
                   for i in range(len(HEADERS))}
            if not rec.get("Company") and not rec.get("Role"):
                continue
            rows.append(rec)
    except Exception:
        return rows
    return rows


def compute_summary(rows):
    """Return dict with Total/Active/Interviews/Offers/Rejected counts."""
    total = len(rows)
    active = interviews = offers = rejected = 0
    interview_statuses = {"Interview Scheduled", "Interviewed",
                          "Technical Test", "Phone Screen"}
    for r in rows:
        status = (r.get("Status") or "").strip()
        if status not in ACTIVE_EXCLUDED:
            active += 1
        if status in interview_statuses:
            interviews += 1
        if status == "Offer":
            offers += 1
        if status == "Rejected":
            rejected += 1
    return {"Total": total, "Active": active, "Interviews": interviews,
            "Offers": offers, "Rejected": rejected}


def export_csv(output_folder):
    """Save a CSV copy of the tracker; return the written path."""
    rows = read_all_rows(output_folder)
    date = datetime.now().strftime("%Y-%m-%d")
    csv_path = os.path.join(output_folder, f"job_applications_{date}.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow({h: r.get(h, "") for h in HEADERS})
    return csv_path
